#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
quality Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'muessise_db',
    'user': 'muessise_admin',
    'password': 'Admin123!'
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (quality sxemi)
# ============================================================

TABLES = {
    'quality_standard': 'Keyfiyyet Standartlari',
    'quality_control_plan': 'Nezaret Planlari',
    'quality_inspection': 'Yoxlamalar',
    'quality_inspection_item': 'Yoxlama Maddeleri',
    'quality_certificate': 'Sertifikatlar',
    'quality_test': 'Testler',
    'quality_non_conformance': 'Uygunsuzluqlar',
    'quality_audit': 'Auditler'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = display_name[:31]
    
    # Başlıq stilləri
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Məlumat stilləri
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sərhəd stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sütun genişlikləri
    column_widths = {
        'id': 8,
        'company_id': 15,
        'code': 15,
        'name': 30,
        'standard_type': 20,
        'description': 40,
        'applicable_products': 35,
        'is_active': 12,
        'created_by': 18,
        'created_at': 22,
        'updated_at': 22,
        'product_id': 18,
        'standard_id': 18,
        'plan_number': 20,
        'control_frequency': 22,
        'sample_size': 18,
        'acceptance_criteria': 35,
        'status': 15,
        'control_plan_id': 22,
        'inspection_number': 20,
        'inspection_date': 18,
        'inspector_id': 18,
        'result': 15,
        'report': 40,
        'notes': 35,
        'inspection_id': 22,
        'parameter_name': 25,
        'parameter_value': 20,
        'min_value': 18,
        'max_value': 18,
        'unit': 12,
        'is_pass': 12,
        'certificate_number': 22,
        'certificate_type': 20,
        'issue_date': 18,
        'expiry_date': 18,
        'issuing_body': 30,
        'test_number': 20,
        'test_date': 18,
        'test_type': 20,
        'test_method': 30,
        'test_result': 18,
        'result_value': 20,
        'laboratory': 30,
        'technician': 25,
        'nc_number': 20,
        'nc_date': 18,
        'severity': 18,
        'root_cause': 40,
        'corrective_action': 40,
        'preventive_action': 40,
        'closed_date': 18,
        'responsible_person': 25,
        'audit_number': 20,
        'audit_date': 18,
        'audit_type': 18,
        'auditor_name': 25,
        'audit_scope': 35,
        'findings': 40,
        'non_conformances': 22,
        'recommendations': 40
    }
    
    # Başlıqları yaz
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Məlumatları yaz
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA
# ============================================================

def main():
    # Çıxış qovluğunu yarat
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Quality')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # PostgreSQL-ə qoşul
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        return
    
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            query = f"SELECT * FROM quality.{table_name} ORDER BY id;"
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

if __name__ == "__main__":
    main()
