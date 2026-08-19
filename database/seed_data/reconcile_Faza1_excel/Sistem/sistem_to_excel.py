#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
sys Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'muessise_db',
    'user': 'muessise_admin',
    'password': 'Admin123!'
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (sys sxemi)
# ============================================================

TABLES = {
    'user': 'Istifadeciler',
    'user_role': 'Istifadeci Rollari',
    'permission': 'Icazeler',
    'role_permission': 'Rol Icazeleri',
    'audit_log': 'Audit Jurnali',
    'system_config': 'Sistem Konfiqurasiyasi',
    'backup_log': 'Backup Qeydleri',
    'error_log': 'Xeta Jurnali',
    'login_history': 'Giris Tarixcesi',
    'notification': 'Bildirislər'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = display_name[:31]
    
    # Başlıq stilləri
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Məlumat stilləri
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sərhəd stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sütun genişlikləri
    column_widths = {
        'id': 8,
        'company_id': 15,
        'employee_id': 18,
        'username': 20,
        'password_hash': 30,
        'email': 30,
        'first_name': 18,
        'last_name': 18,
        'is_active': 12,
        'is_locked': 12,
        'last_login': 22,
        'created_at': 22,
        'updated_at': 22,
        'role_code': 18,
        'role_name': 20,
        'description': 40,
        'priority': 12,
        'permission_code': 25,
        'permission_name': 25,
        'module_name': 20,
        'role_id': 18,
        'permission_id': 18,
        'action': 20,
        'table_name': 20,
        'record_id': 15,
        'old_value': 35,
        'new_value': 35,
        'ip_address': 18,
        'user_agent': 35,
        'config_key': 25,
        'config_value': 30,
        'config_type': 18,
        'backup_number': 20,
        'backup_date': 22,
        'backup_type': 18,
        'backup_size': 18,
        'backup_location': 40,
        'status': 15,
        'notes': 35,
        'created_by': 18,
        'error_code': 18,
        'error_message': 40,
        'error_level': 18,
        'file_name': 25,
        'line_number': 15,
        'stack_trace': 40,
        'user_id': 18,
        'login_time': 22,
        'logout_time': 22,
        'session_id': 25,
        'notification_type': 20,
        'title': 30,
        'message': 40,
        'link': 35,
        'is_read': 12,
        'user_agent': 35,
        'session_id': 30,
        'user_id': 18
    }
    
    # Başlıqları yaz
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Məlumatları yaz
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA
# ============================================================

def main():
    # Çıxış qovluğunu yarat
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Sistem')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # PostgreSQL-ə qoşul
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        return
    
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            query = f"SELECT * FROM sys.{table_name} ORDER BY id;"
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

if __name__ == "__main__":
    main()
