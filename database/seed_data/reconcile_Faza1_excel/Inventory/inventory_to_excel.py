#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
inventory Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'muessise_db',
    'user': 'muessise_admin',
    'password': 'Admin123!'
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (inventory sxemi)
# ============================================================

TABLES = {
    'warehouse': 'Anbar',
    'warehouse_zone': 'Anbar Zonalari',
    'product_stock': 'Mehsul Stoku',
    'stock_movement': 'Stok Hereketleri',
    'inventory_count': 'Inventarizasiya',
    'inventory_count_item': 'Inventarizasiya Maddeleri',
    'stock_transfer': 'Stok Transferi',
    'stock_adjustment': 'Stok Duzelishi',
    'warehouse_location': 'Anbar Yerleri'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = display_name[:31]
    
    # Başlıq stilləri
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Məlumat stilləri
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sərhəd stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sütun genişlikləri
    column_widths = {
        'id': 8,
        'company_id': 15,
        'branch_id': 15,
        'code': 15,
        'name': 30,
        'address': 40,
        'warehouse_type': 20,
        'capacity': 15,
        'is_active': 12,
        'created_at': 22,
        'updated_at': 22,
        'warehouse_id': 18,
        'description': 40,
        'product_id': 18,
        'warehouse_zone_id': 22,
        'quantity': 18,
        'reserved_quantity': 22,
        'min_stock': 18,
        'max_stock': 18,
        'last_updated': 22,
        'movement_type': 18,
        'movement_date': 22,
        'reference_type': 20,
        'reference_id': 18,
        'created_by': 18,
        'count_number': 20,
        'count_date': 18,
        'count_type': 18,
        'status': 15,
        'counted_by': 18,
        'approved_by': 18,
        'count_id': 18,
        'system_quantity': 20,
        'counted_quantity': 20,
        'difference': 18,
        'notes': 35,
        'transfer_number': 20,
        'transfer_date': 18,
        'from_warehouse_id': 22,
        'to_warehouse_id': 20,
        'adjustment_number': 20,
        'adjustment_date': 18,
        'adjustment_type': 20,
        'previous_quantity': 22,
        'new_quantity': 18,
        'reason': 40,
        'zone_id': 18,
        'location_type': 18
    }
    
    # Başlıqları yaz
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Məlumatları yaz
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA
# ============================================================

def main():
    # Çıxış qovluğunu yarat
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Inventory')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # PostgreSQL-ə qoşul
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        return
    
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            query = f"SELECT * FROM inventory.{table_name} ORDER BY id;"
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

if __name__ == "__main__":
    main()
