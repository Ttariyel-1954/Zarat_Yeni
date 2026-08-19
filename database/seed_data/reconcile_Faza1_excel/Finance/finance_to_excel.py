#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
finance Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'muessise_db',
    'user': 'muessise_admin',
    'password': 'Admin123!'
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (finance sxemi)
# ============================================================

TABLES = {
    'chart_of_accounts': 'Hesablar Plani',
    'journal_entry': 'Jurnal Yazilislari',
    'journal_entry_line': 'Jurnal Setirleri',
    'general_ledger': 'Bas Kitab',
    'accounts_receivable': 'Debitor Borclari',
    'accounts_payable': 'Kreditor Borclari',
    'bank_account': 'Bank Hesablari',
    'bank_transaction': 'Bank Emeliyyatlari',
    'budget': 'Budceler',
    'budget_item': 'Budce Maddeleri',
    'tax': 'Vergiler',
    'tax_return': 'Vergi Beyannameleri',
    'financial_report': 'Maliyye Hesabatlari'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = display_name[:31]
    
    # Başlıq stilləri
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Məlumat stilləri
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sərhəd stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sütun genişlikləri
    column_widths = {
        'id': 8,
        'company_id': 15,
        'account_code': 18,
        'account_name': 30,
        'account_type': 18,
        'parent_account_id': 22,
        'normal_balance': 18,
        'description': 40,
        'is_active': 12,
        'created_at': 22,
        'updated_at': 22,
        'entry_number': 20,
        'entry_date': 18,
        'reference_type': 22,
        'reference_id': 18,
        'total_debit': 18,
        'total_credit': 18,
        'status': 15,
        'created_by': 18,
        'approved_by': 18,
        'entry_id': 18,
        'account_id': 18,
        'debit_amount': 18,
        'credit_amount': 18,
        'balance': 18,
        'customer_id': 18,
        'invoice_id': 18,
        'receivable_number': 22,
        'transaction_date': 18,
        'due_date': 18,
        'total_amount': 18,
        'paid_amount': 18,
        'remaining_amount': 18,
        'notes': 35,
        'supplier_id': 18,
        'payable_number': 22,
        'account_number': 22,
        'bank_name': 25,
        'branch_code': 18,
        'iban': 30,
        'currency': 12,
        'account_type': 18,
        'bank_account_id': 22,
        'transaction_number': 22,
        'transaction_type': 20,
        'amount': 18,
        'counterparty': 25,
        'reference': 20,
        'budget_number': 20,
        'budget_name': 30,
        'budget_year': 15,
        'budget_type': 18,
        'budget_id': 18,
        'planned_amount': 18,
        'actual_amount': 18,
        'variance': 18,
        'tax_code': 18,
        'tax_name': 25,
        'tax_type': 18,
        'tax_rate': 15,
        'applicable_from': 18,
        'applicable_to': 18,
        'return_number': 20,
        'return_period': 18,
        'period_start': 18,
        'period_end': 18,
        'submission_date': 18,
        'tax_amount': 18,
        'report_number': 20,
        'report_name': 30,
        'report_type': 22,
        'report_period': 18,
        'report_data': 35,
        'total_assets': 18,
        'total_liabilities': 22,
        'total_equity': 18,
        'total_revenue': 18,
        'total_expenses': 18,
        'net_profit': 18,
        'tax_id': 15
    }
    
    # Başlıqları yaz
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Məlumatları yaz
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA
# ============================================================

def main():
    # Çıxış qovluğunu yarat
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Finance')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # PostgreSQL-ə qoşul
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        return
    
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            query = f"SELECT * FROM finance.{table_name} ORDER BY id;"
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

if __name__ == "__main__":
    main()
