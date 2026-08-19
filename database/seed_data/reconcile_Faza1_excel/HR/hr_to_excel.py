#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
hr Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',      # PostgreSQL serverinin ünvanı (lokal kompüter)
    'port': 5433,             # PostgreSQL 18-in işlədiyi port
    'database': 'muessise_db', # Verilənlər bazasının adı
    'user': 'muessise_admin',  # İstifadəçi adı
    'password': 'Admin123!'    # İstifadəçi şifrəsi
}

def get_connection():
    """PostgreSQL-ə qoşulur"""
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (hr sxemi)
# ============================================================

TABLES = {
    'employee_detail': 'Isci Etrafli Melumatlari',
    'payroll': 'Emek Haqqi Melumatlari',
    'salary_history': 'Maas Tarixcesi',
    'attendance': 'Ise Gelme Getme Qeydleri',
    'vacation': 'Mezuniyyet Melumatlari',
    'sick_leave': 'Xestelik Vereqeleri',
    'bonus': 'Mukafat Bonus Melumatlari',
    'deduction': 'Tutulmalar',
    'training': 'Telim Melumatlari',
    'performance_review': 'Performans Qiymetlendirmesi',
    'termination': 'Isden Cixarilma Melumatlari'
}
# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    """
    Cədvəl məlumatlarından Excel faylı yaradır
    """
    # DataFrame yarat (pandas cədvəli)
    df = pd.DataFrame(data)
    
    # Excel faylı yarat
    wb = Workbook()
    ws = wb.active
    ws.title = display_name
    
    # ========== BAŞLIQ STİLLƏRİ ==========
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ========== MƏLUMAT STİLLƏRİ ==========
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # ========== SƏRHƏD STİLLƏRİ ==========
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== SÜTUN GENİŞLİKLƏRİ ==========
    column_widths = {
        'id': 8,
        'employee_id': 15,
        'fin_code': 15,
        'passport_series': 18,
        'passport_number': 20,
        'address': 35,
        'city': 18,
        'district': 18,
        'education': 20,
        'specialty': 25,
        'marital_status': 18,
        'children_count': 15,
        'emergency_contact_name': 25,
        'emergency_contact_phone': 22,
        'bank_name': 20,
        'bank_account': 25,
        'created_at': 22,
        'updated_at': 22,
        'payroll_month': 18,
        'base_salary': 18,
        'bonus_amount': 18,
        'allowance_amount': 20,
        'overtime_amount': 18,
        'gross_salary': 18,
        'tax_amount': 18,
        'social_security': 20,
        'deduction_amount': 18,
        'net_salary': 18,
        'payment_date': 18,
        'payment_status': 20,
        'previous_salary': 18,
        'new_salary': 18,
        'change_reason': 30,
        'change_date': 18,
        'approved_by': 15,
        'attendance_date': 18,
        'check_in_time': 15,
        'check_out_time': 15,
        'work_hours': 15,
        'is_late': 12,
        'is_overtime': 15,
        'status': 15,
        'vacation_type': 20,
        'start_date': 18,
        'end_date': 18,
        'total_days': 15,
        'used_days': 15,
        'remaining_days': 18,
        'sick_type': 20,
        'doctor_name': 25,
        'hospital_name': 25,
        'certificate_number': 22,
        'bonus_type': 20,
        'amount': 18,
        'reason': 35,
        'bonus_date': 18,
        'deduction_type': 20,
        'deduction_date': 18,
        'training_name': 30,
        'training_type': 20,
        'total_hours': 15,
        'grade': 10,
        'certificate': 15,
        'certificate_number': 22,
        'review_date': 18,
        'review_type': 20,
        'rating': 12,
        'strengths': 35,
        'weaknesses': 35,
        'goals': 35,
        'reviewer_id': 15,
        'termination_date': 18,
        'termination_type': 20,
        'compensation_amount': 22,
        'notice_period_days': 20,
        'recommendation': 35
    }
    
    # ========== BAŞLIQLARI YAZ ==========
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Sütun genişliyini təyin et
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # ========== MƏLUMATLARI YAZ ==========
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # ========== FAYLI SAXLA ==========
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA (proqramın başlanğıc nöqtəsi)
# ============================================================

def main():
    """Bütün cədvəlləri Excel-ə çevirir"""
    
    # ===== Çıxış qovluğunu yarat =====
    # Faylların saxlanacağı yer: ~/Desktop/postgres_dersleri/HR/
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'HR')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # ===== PostgreSQL-ə qoşul =====
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        print("\nYoxlayın:")
        print("1. PostgreSQL 18 işləyir? (brew services list)")
        print("2. Port düzgündür? (5433)")
        print("3. İstifadəçi və şifrə düzgündür?")
        return
    
    # ===== Hər bir cədvəl üçün =====
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            # SQL sorğusu (bütün məlumatları seçir)
            query = f"SELECT * FROM hr.{table_name} ORDER BY id;"
            
            # Məlumatları PostgreSQL-dən oxu
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            # Excel faylı yarat
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    # ===== Bağla =====
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

# ============================================================
# PROQRAMI İŞƏ SAL
# ============================================================

if __name__ == "__main__":
    main()
