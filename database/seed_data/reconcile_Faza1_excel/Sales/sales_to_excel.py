#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
sales Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'muessise_db',
    'user': 'muessise_admin',
    'password': 'Admin123!'
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (sales sxemi)
# ============================================================

TABLES = {
    'customer': 'Musteri Melumatlari',
    'customer_address': 'Musteri Unvanlari',
    'sales_order': 'Satis Sifarisleri',
    'sales_order_item': 'Sifaris Maddeleri',
    'delivery_order': 'Catdirilma Sifarisleri',
    'delivery_order_item': 'Catdirilma Maddeleri',
    'sales_invoice': 'Satis Hesab Fakturalari',
    'sales_return': 'Satis Qaytarmalari'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = display_name[:31]
    
    # Başlıq stilləri
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Məlumat stilləri
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sərhəd stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sütun genişlikləri
    column_widths = {
        'id': 8,
        'company_id': 15,
        'code': 15,
        'name': 30,
        'tax_id': 18,
        'phone': 18,
        'email': 30,
        'website': 25,
        'customer_type': 20,
        'credit_limit': 18,
        'payment_term_days': 20,
        'is_active': 12,
        'created_at': 22,
        'updated_at': 22,
        'customer_id': 18,
        'address_type': 18,
        'address_line1': 35,
        'address_line2': 35,
        'city': 20,
        'district': 20,
        'postal_code': 15,
        'country': 18,
        'is_primary': 15,
        'order_number': 20,
        'order_date': 18,
        'delivery_date': 18,
        'status': 15,
        'total_amount': 18,
        'tax_amount': 18,
        'grand_total': 18,
        'description': 40,
        'created_by': 18,
        'approved_by': 18,
        'order_id': 18,
        'product_id': 18,
        'quantity': 18,
        'unit_price': 18,
        'total_price': 18,
        'discount_percent': 20,
        'discount_amount': 20,
        'tax_rate': 15,
        'line_total': 18,
        'sales_order_id': 22,
        'delivery_number': 20,
        'shipping_method': 25,
        'tracking_number': 22,
        'notes': 35,
        'delivery_id': 18,
        'delivered_quantity': 22,
        'invoice_number': 20,
        'invoice_date': 18,
        'due_date': 18,
        'payment_status': 18,
        'payment_date': 18,
        'invoice_id': 18,
        'return_number': 20,
        'return_date': 18,
        'reason': 25,
        'customer_id': 18
    }
    
    # Başlıqları yaz
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Məlumatları yaz
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA
# ============================================================

def main():
    # Çıxış qovluğunu yarat
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Sales')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # PostgreSQL-ə qoşul
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        return
    
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            query = f"SELECT * FROM sales.{table_name} ORDER BY id;"
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

if __name__ == "__main__":
    main()
