#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
logistics Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',
    'port': 5433,
    'database': 'muessise_db',
    'user': 'muessise_admin',
    'password': 'Admin123!'
}

def get_connection():
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (logistics sxemi)
# ============================================================

TABLES = {
    'vehicle': 'Neqliyyat Vasiteleri',
    'driver': 'Suruculer',
    'route': 'Marsrutlar',
    'shipment': 'Dasimalar',
    'shipment_item': 'Dasima Maddeleri',
    'delivery_tracking': 'Catdirilma Izleme'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    df = pd.DataFrame(data)
    
    wb = Workbook()
    ws = wb.active
    ws.title = display_name[:31]
    
    # Başlıq stilləri
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # Məlumat stilləri
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Sərhəd stilləri
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sütun genişlikləri
    column_widths = {
        'id': 8,
        'company_id': 15,
        'vehicle_code': 18,
        'plate_number': 15,
        'brand': 20,
        'model': 20,
        'year': 12,
        'capacity_kg': 18,
        'fuel_type': 15,
        'status': 15,
        'last_maintenance_date': 25,
        'is_active': 12,
        'created_at': 22,
        'updated_at': 22,
        'first_name': 18,
        'last_name': 18,
        'phone': 18,
        'email': 30,
        'license_number': 22,
        'license_class': 18,
        'license_expiry': 18,
        'hire_date': 18,
        'notes': 35,
        'route_code': 18,
        'origin': 20,
        'destination': 20,
        'distance_km': 18,
        'estimated_hours': 18,
        'stops': 35,
        'shipment_number': 22,
        'route_id': 18,
        'vehicle_id': 18,
        'driver_id': 18,
        'shipment_date': 18,
        'departure_time': 15,
        'arrival_time': 15,
        'total_weight_kg': 22,
        'total_volume_m3': 20,
        'created_by': 18,
        'shipment_id': 18,
        'product_id': 18,
        'quantity': 18,
        'weight_kg': 18,
        'volume_m3': 18,
        'tracking_number': 25,
        'update_time': 22,
        'location': 25,
        'latitude': 15,
        'longitude': 15,
        'driver_status': 15,
        'shipment_status': 18,
        'vehicle_status': 18
    }
    
    # Başlıqları yaz
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        col_width = column_widths.get(col_name, 20)
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # Məlumatları yaz
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA
# ============================================================

def main():
    # Çıxış qovluğunu yarat
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Logistics')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # PostgreSQL-ə qoşul
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        return
    
    success_count = 0
    for table_name, display_name in TABLES.items():
        try:
            query = f"SELECT * FROM logistics.{table_name} ORDER BY id;"
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            create_excel_file(table_name, display_name, df, output_dir)
            success_count += 1
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {success_count} Excel faylı yaradıldı.")

if __name__ == "__main__":
    main()
