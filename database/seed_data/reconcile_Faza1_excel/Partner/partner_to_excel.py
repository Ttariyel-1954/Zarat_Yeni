#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
partner Sxemi - Excel Fayllarının Yaradılması
PostgreSQL-dən məlumatları oxuyub .xlsx formatında saxlayır
"""

import psycopg2
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import os

# ============================================================
# VERİLƏNLƏR BAZASINA QOŞULMA MƏLUMATLARI
# ============================================================

DB_CONFIG = {
    'host': 'localhost',      # PostgreSQL serverinin ünvanı (lokal kompüter)
    'port': 5433,             # PostgreSQL 18-in işlədiyi port
    'database': 'muessise_db', # Verilənlər bazasının adı
    'user': 'muessise_admin',  # İstifadəçi adı
    'password': 'Admin123!'    # İstifadəçi şifrəsi
}

def get_connection():
    """PostgreSQL-ə qoşulur"""
    return psycopg2.connect(**DB_CONFIG)

# ============================================================
# CƏDVƏL SİYAHISI (partner sxemi)
# ============================================================

TABLES = {
    'partner_type': 'Tərəfdaş Növləri',
    'partner_category': 'Tərəfdaş Kateqoriyaları',
    'partner': 'Əsas Tərəfdaş Məlumatları',
    'contact_person': 'Əlaqə Şəxsləri',
    'address': 'Ünvan Məlumatları',
    'bank_account': 'Bank Hesabları',
    'contract': 'Müqavilələr',
    'partner_rating': 'Tərəfdaş Reytinqi',
    'partner_activity_log': 'Fəaliyyət Jurnalı'
}

# ============================================================
# EXCEL FAYLI YARADAN FUNKSİYA
# ============================================================

def create_excel_file(table_name, display_name, data, output_dir):
    """
    Cədvəl məlumatlarından Excel faylı yaradır
    """
    # DataFrame yarat (pandas cədvəli)
    df = pd.DataFrame(data)
    
    # Excel faylı yarat
    wb = Workbook()
    ws = wb.active
    ws.title = display_name
    
    # ========== BAŞLIQ STİLLƏRİ ==========
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4F46E5', end_color='4F46E5', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # ========== MƏLUMAT STİLLƏRİ ==========
    data_font = Font(name='Arial', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # ========== SƏRHƏD STİLLƏRİ ==========
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== SÜTUN GENİŞLİKLƏRİ ==========
    column_widths = {
        'id': 8,
        'code': 15,
        'name': 30,
        'description': 40,
        'is_active': 12,
        'created_at': 22,
        'updated_at': 22,
        'company_id': 15,
        'partner_type_id': 18,
        'partner_category_id': 22,
        'tax_id': 18,
        'registration_number': 22,
        'phone': 18,
        'email': 30,
        'website': 25,
        'credit_limit': 18,
        'payment_term_days': 20,
        'first_name': 18,
        'last_name': 18,
        'position': 25,
        'department': 20,
        'mobile': 18,
        'is_primary': 15,
        'address_type': 18,
        'address_line1': 35,
        'address_line2': 35,
        'city': 20,
        'district': 20,
        'postal_code': 15,
        'country': 18,
        'bank_name': 25,
        'bank_code': 18,
        'account_number': 22,
        'iban': 30,
        'currency': 12,
        'account_type': 18,
        'is_active': 12,
        'contract_number': 22,
        'contract_type': 20,
        'start_date': 18,
        'end_date': 18,
        'contract_amount': 20,
        'status': 15,
        'signed_by': 15,
        'priority': 12,
        'rating_date': 18,
        'reliability_score': 20,
        'quality_score': 18,
        'payment_score': 18,
        'delivery_score': 18,
        'overall_rating': 18,
        'review_text': 40,
        'reviewed_by': 15,
        'activity_type': 20,
        'activity_date': 22,
        'contact_person_id': 22,
        'performed_by': 18
    }
    
    # ========== BAŞLIQLARI YAZ ==========
    for col_idx, col_name in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        
        # Sütun genişliyini təyin et
        col_width = column_widths.get(col_name, 20)
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(col_idx)].width = col_width
    
    # ========== MƏLUMATLARI YAZ ==========
    for row_idx, row in enumerate(df.values, 2):
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = thin_border
    
    # ========== FAYLI SAXLA ==========
    filename = f"{table_name}.xlsx"
    filepath = os.path.join(output_dir, filename)
    wb.save(filepath)
    print(f"✅ {filename} - {len(df)} sətir məlumatla yaradıldı")
    return filepath

# ============================================================
# ƏSAS FONKSİYA (proqramın başlanğıc nöqtəsi)
# ============================================================

def main():
    """Bütün cədvəlləri Excel-ə çevirir"""
    
    # ===== Çıxış qovluğunu yarat =====
    # Faylların saxlanacağı yer: ~/Desktop/postgres_dersleri/Partner/
    output_dir = os.path.join(os.path.expanduser('~/Desktop'), 'postgres_dersleri', 'Partner')
    os.makedirs(output_dir, exist_ok=True)
    print(f"📁 Çıxış qovluğu: {output_dir}\n")
    
    # ===== PostgreSQL-ə qoşul =====
    try:
        conn = get_connection()
        print("✅ PostgreSQL-ə qoşuldu\n")
    except Exception as e:
        print(f"❌ Qoşulma xətası: {e}")
        print("\nYoxlayın:")
        print("1. PostgreSQL 18 işləyir? (brew services list)")
        print("2. Port düzgündür? (5433)")
        print("3. İstifadəçi və şifrə düzgündür?")
        return
    
    # ===== Hər bir cədvəl üçün =====
    for table_name, display_name in TABLES.items():
        try:
            # SQL sorğusu (bütün məlumatları seçir)
            query = f"SELECT * FROM partner.{table_name} ORDER BY id;"
            
            # Məlumatları PostgreSQL-dən oxu
            df = pd.read_sql(query, conn)
            
            if df.empty:
                print(f"⚠️ {table_name} - məlumat yoxdur")
                continue
            
            # Excel faylı yarat
            create_excel_file(table_name, display_name, df, output_dir)
            
        except Exception as e:
            print(f"❌ {table_name} xətası: {e}")
    
    # ===== Bağla =====
    conn.close()
    print(f"\n🎉 Bütün fayllar {output_dir} qovluğuna yaradıldı!")
    print(f"📊 Cəmi {len(TABLES)} Excel faylı yaradıldı.")

# ============================================================
# PROQRAMI İŞƏ SAL
# ============================================================

if __name__ == "__main__":
    main()
